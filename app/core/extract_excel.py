from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook


def extract_text_from_excel(path: Path) -> Iterator[str]:
    """Yield text lines from an Excel file by flattening sheets with coordinates using openpyxl."""
    wb = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                pieces = []
                for idx, value in enumerate(row, start=1):
                    if value is None:
                        continue
                    value_str = str(value).strip()
                    if value_str == "":
                        continue
                    pieces.append(f"C{idx}:{value_str}")
                if pieces:
                    yield f"{sheet_name} | " + " | ".join(pieces)
    finally:
        wb.close()